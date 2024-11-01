import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import os
import openpyxl
import numpy as np
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def scrape_exchange_rates(currency, start_date):
    url = f"https://www.x-rates.com/historical/?from={currency[0]}&to={currency[1]}&amount=1&date={start_date}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    rates = {}
    for row in soup.find_all('tr')[1:]:
        cols = row.find_all('td')
        if len(cols) > 1:
            date = cols[0].text.strip()
            rate = cols[1].text.strip()
            time = cols[2].text.strip() if len(cols) > 2 else ''
            rates[date] = {'rate': rate, 'time': time}

    print(f"Scraped rates for {currency[0]} to {currency[1]}: {rates}")
    return rates

def prepare_data(rates, currency):
    df = pd.DataFrame(list(rates.items()), columns=['Дата', 'Данные'])
    df['Курс'] = df['Данные'].apply(lambda x: pd.to_numeric(x['rate'].replace(',', ''), errors='coerce'))
    df['Время'] = df['Данные'].apply(lambda x: x['time'])
    df = df[['Дата', 'Курс', 'Время']]
    df.columns = [f'{currency[0]}/{currency[1]} {col}' for col in df.columns]
    print(f"Prepared DataFrame for {currency[0]}: {df}")
    return df

def save_to_excel(df_usd, df_jpy):
    file_exists = os.path.exists('exchange_rates.xlsx')
    
    with pd.ExcelWriter('exchange_rates.xlsx', engine='openpyxl', mode='a' if file_exists else 'w') as writer:
        if not file_exists:
            df_usd.to_excel(writer, sheet_name='Sheet1', index=False, startcol=0)
            df_jpy.to_excel(writer, sheet_name='Sheet1', index=False, startcol=3)
        else:
            try:
                writer.book = openpyxl.load_workbook('exchange_rates.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                
                df_usd.to_excel(writer, sheet_name='Sheet1', index=False, startcol=0, header=False)
                df_jpy.to_excel(writer, sheet_name='Sheet1', index=False, startcol=3, header=False)
            except Exception as e:
                print(f"Ошибка при загрузке файла Excel: {e}")
                return

def calculate_result(df_usd, df_jpy):
    usd_to_rub = df_usd[f'USD/RUB Курс'].astype(float)
    jpy_to_rub = df_jpy[f'JPY/RUB Курс'].astype(float)

    print("USD to RUB rates:")
    print(usd_to_rub)
    print("JPY to RUB rates:")
    print(jpy_to_rub)

    # рассчет результата, игнорируя NaN значения
    result = usd_to_rub / jpy_to_rub
    result = result.replace([np.inf, -np.inf], np.nan)  # заменяем бесконечности на NaN
    result = result.dropna()  # удаляем строки с NaN значениями

    print("Calculated results:")
    print(result)

    df_result = pd.DataFrame({
        'Дата': df_usd[f'USD/RUB Дата'].iloc[:len(result)],
        'Результат': result
    })
    return df_result

def write_result_to_excel(df_result):
    workbook = openpyxl.load_workbook('exchange_rates.xlsx')
    sheet = workbook['Sheet1']
    
    start_row = 2
    while sheet.cell(row=start_row, column=7).value:
        start_row += 1

    for i, (date, result) in enumerate(zip(df_result['Дата'], df_result['Результат']), start=start_row):
        print(f"Writing to row {i}: Result: {result}, Date: {date}")
        sheet.cell(row=i, column=7, value=result)
        sheet.cell(row=i, column=8, value=date)

    workbook.save('exchange_rates.xlsx')

def send_email(filename, row_count): #свои реальные данные вводить не стал
    sender_email = "email@example.com"
    receiver_email = "emailrecieve@example.com"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = "Отчет по курсам валют"

    if row_count % 10 == 1 and row_count % 100 != 11:
        string_form = "строка"
    elif 2 <= row_count % 10 <= 4 and not (12 <= row_count % 100 <= 14):
        string_form = "строки"
    else:
        string_form = "строк"

    body = f"В отчете содержится {row_count} {string_form}."
    msg.attach(MIMEText(body, 'plain'))

    with open(filename, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)

    with smtplib.SMTP('examplesmtp.com', 587) as server: #ввести реальный адрес сервера
        server.starttls()
        server.login(sender_email, password)
        server.send_message(msg)

if __name__ == "__main__":
    end_date = datetime.now() - timedelta(days=1)
    start_date = end_date - timedelta(days=30)
    start_date_str = start_date.strftime('%Y-%m-%d')
    
    rates_usd = scrape_exchange_rates(['USD', 'RUB'], start_date_str)
    rates_jpy = scrape_exchange_rates(['JPY', 'RUB'], start_date_str)

    df_usd = prepare_data(rates_usd, ['USD', 'RUB'])
    df_jpy = prepare_data(rates_jpy, ['JPY', 'RUB'])

    save_to_excel(df_usd, df_jpy)

    df_result = calculate_result(df_usd, df_jpy)
    write_result_to_excel(df_result)

    row_count = df_usd.shape[0]
    send_email('exchange_rates.xlsx', row_count)
